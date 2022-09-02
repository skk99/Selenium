import openpyxl

# Excel File contains workbook-->sheets-->rows-->cells
# Get the file, Load the workbook and then sheet
file = r"C:\Users\shakarn\Documents\data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

# Find rows and cols
rows = sheet.max_row        # Get the no of rows in an excel sheet
cols = sheet.max_column        # Get the no of cols in an excel sheet


# Reading all the rows and cols from excel sheet
for row in range(1, rows+1):
    for col in range(1, cols+1):
        print(sheet.cell(row, col).value, end = '            ')           # Get the cell value at the given row and col
    print()

