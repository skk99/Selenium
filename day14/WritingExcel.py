import openpyxl

# file = r"C:\Users\shakarn\Documents\testdata.xlsx"
# workbook = openpyxl.load_workbook(file)
# # sheet = workbook["Sheet1"]
# # sheet = workbook["Data"]        # If we have a customized sheet name (use if we have multiple sheets)
# sheet = workbook.active         # Get active sheet from excel if we have only one sheet in excel
#
#
# # Design how many rows and cols you want to write
# for r in range(1, 6):           # 5 rows
#     for c in range(1, 4):       # 3 cols
#         sheet.cell(r, c).value = "welcome"          # write the data in the given r and c cell
#
# # save the file
# workbook.save(file)

# Now, To write different data in each rows and cols
file = r"C:\Users\shakarn\Documents\testdata1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1, 1).value = 123        #1st row 1st col
sheet.cell(1,2).value = "smith"     #1st row 2nd col
sheet.cell(1,3).value = "engineer"  #1st row 3rd col

sheet.cell(2, 1).value = 456        #2nd row 1st col
sheet.cell(2,2).value = "jon"     #2nd row 2nd col
sheet.cell(2,3).value = "doctor"  #2nd row 3rd col

sheet.cell(3, 1).value = 789        #3rd row 1st col
sheet.cell(3,2).value = "james"     #3rd row 2nd col
sheet.cell(3,3).value = "lawyer"    #3rd row 3rd col

workbook.save(file)                 #save the file